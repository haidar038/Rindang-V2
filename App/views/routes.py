from flask import Blueprint, current_app, request, render_template, flash, redirect, url_for, jsonify, session
from flask_login import login_required, current_user
from flask_sqlalchemy import pagination
from flask_admin.base import expose, AdminIndexView, Admin
from werkzeug.security import check_password_hash, generate_password_hash
from sqlalchemy import asc
from datetime import datetime, timedelta
from babel.numbers import format_currency
from werkzeug.utils import secure_filename
import json, requests, secrets, os

from App.models import User, DataPangan, Kelurahan
from App import db, UPLOAD_FOLDER

views = Blueprint('views', __name__)

PICTURE_ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
REPORT_ALLOWED_EXTENSIONS = {'xlsx'}
REPORT_STAT = {'panen', 'penanaman'}

def picture_allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in PICTURE_ALLOWED_EXTENSIONS

# Fungsi yang disederhanakan untuk memeriksa nama file dan tipe impor
def allowed_report_stat(filename, import_type):
    file_name_without_extension = os.path.splitext(filename)[0].lower()
    return import_type in file_name_without_extension

def report_allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in REPORT_ALLOWED_EXTENSIONS

# Constants for API URL parameters
KAB_KOTA = 458  # Ternate
KOMODITAS_ID = 3
TARGET_KOMODITAS = ["Cabai Merah Keriting", "Cabai Rawit Merah", "Bawang Merah"]

# Helper function to fetch and format price data from API
def fetch_price_data(start_date, end_date):
    """Fetches price data from the API and formats it for display.

    Args:
        start_date (str): The starting date in YYYY-MM-DD format.
        end_date (str): The ending date in YYYY-MM-DD format.

    Returns:
        list: A list of dictionaries containing formatted price data.
    """

    url = f"https://panelharga.badanpangan.go.id/data/kabkota-range-by-levelharga/{KAB_KOTA}/{KOMODITAS_ID}/{start_date}/{end_date}"

    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for HTTP errors

        data = response.json()

        table_data = []
        for item in data["data"]:
            if item["name"] in TARGET_KOMODITAS:
                for date_data in item["by_date"]:
                    date_obj = datetime.strptime(date_data["date"], "%Y-%m-%d")
                    formatted_date = date_obj.strftime("%d/%m/%Y")
                    geomean_value = date_data["geomean"]

                    # Menyederhanakan format harga
                    formatted_price = "-" if geomean_value == "-" else format_currency(float(geomean_value), "IDR", locale="id_ID", decimal_quantization=False)[:-3]

                    table_data.append({
                        "date": formatted_date,
                        "name": item["name"],
                        "price": formatted_price
                    })

        return table_data

    except requests.exceptions.RequestException as e:
        flash(f"Error fetching data: {e}", category='error')
        return []  # Return an empty list on error

@views.route('/dashboard/profil/<int:id>/update_picture', methods=['POST'])
@login_required
def update_profile_picture(id):
    upload_folder = current_app.config['UPLOAD_FOLDER']
    user = User.query.get_or_404(id)
    if request.method == 'POST':
        if 'profile_pic' not in request.files:
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)
        file = request.files['profile_pic']
        if file.filename == '':
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)
        if file and picture_allowed_file(file.filename):
            filename = secrets.token_hex(8) + '_' + secure_filename(file.filename)
            file_path = os.path.join(upload_folder, filename)
            file.save(file_path)

            # Hapus foto profil lama jika ada
            if user.profile_pic:
                old_file_path = os.path.join(upload_folder, user.profile_pic)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)

            user.profile_pic = filename
            db.session.commit()
            flash('Foto profil berhasil diubah!', 'success')
        else:
            flash('File yang diizinkan hanya JPG, JPEG, dan PNG.', 'error')
    return redirect(url_for('views.profil'))

@views.route('/', methods=['POST', 'GET'])
def index():
    kelurahan = Kelurahan.query.all()
    produksi = DataPangan.query.all()

    # Menggunakan metode today() untuk mendapatkan tanggal hari ini
    today = datetime.today()
    one_week_ago = today - timedelta(days=7)
    start_date = one_week_ago.strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")

    table_data = fetch_price_data(start_date, end_date)

    total_kebun = sum(kel.kebun for kel in kelurahan)
    total_panen = sum(prod.jml_panen for prod in produksi)

    return render_template('index.html', table_data=table_data, kebun=total_kebun, produksi=total_panen, round=round)

@views.route('/dashboard', methods=['GET', 'POST'])
@login_required
def dashboard():
    if current_user.account_type == 'admin':
        return redirect(url_for('admin_page.index'))

    data_pangan = DataPangan.query.filter_by(user_id=current_user.id).all()

    # Menggunakan metode today() untuk mendapatkan tanggal hari ini
    today = datetime.today()
    one_week_ago = today - timedelta(days=7)
    start_date = one_week_ago.strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")

    table_data = fetch_price_data(start_date, end_date)

    total_panen = sum(prod.jml_panen for prod in data_pangan if data_pangan)

    return render_template('dashboard/index.html', total_panen=total_panen, harga=table_data, round=round)

@views.route('/dashboard/penjualan')
@login_required
def penjualan():
    return render_template('dashboard/penjualan.html')

@views.route('/dashboard/harga-pangan', methods=['POST', 'GET'])
@login_required
def hargapangan():
    if current_user.account_type == 'admin':
        return redirect(url_for('admin_page.index'))

    # Menggunakan metode today() untuk mendapatkan tanggal hari ini
    today = datetime.today()
    one_week_ago = today - timedelta(days=7)
    start_date = one_week_ago.strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")

    table_data = fetch_price_data(start_date, end_date)

    return render_template('dashboard/harga-pangan.html', table_data=table_data)

@views.route('/dashboard/data-pangan', methods=['POST','GET'])
@login_required
def dataproduksi():
    if current_user.account_type == 'admin':
        return redirect(url_for('admin_page.index'))

    user_data = User.query.filter_by(id=current_user.id).first()
    pangan = DataPangan.query.filter_by(user_id=current_user.id).all()
    kel = Kelurahan.query.filter_by(id=current_user.kelurahan_id).first()

    page = request.args.get('page', 1, type=int)  # Get page number from query string
    per_page = 5 # Number of items per page

    # Menggunakan list comprehension untuk menyederhanakan perhitungan total panen
    total_panen = [total.jml_panen for total in pangan]

    cabai = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Cabai').order_by(asc(DataPangan.tanggal_panen)).paginate(page=page, per_page=per_page, error_out=False)
    tomat = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Tomat').order_by(asc(DataPangan.tanggal_panen)).paginate(page=page, per_page=per_page, error_out=False)

    allDataCabai = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Cabai').order_by(asc(DataPangan.tanggal_panen)).all()
    allDataTomat = DataPangan.query.filter_by(user_id=current_user.id, komoditas='Tomat').order_by(asc(DataPangan.tanggal_panen)).all()

    # Menggunakan list comprehension untuk menyederhanakan pembuatan list data statistik
    stat_cabai = [panenCabai.jml_panen for panenCabai in allDataCabai]
    tgl_panen_cabai = [panenCabai.tanggal_panen for panenCabai in allDataCabai]
    stat_tomat = [panenTomat.jml_panen for panenTomat in allDataTomat]
    tgl_panen_tomat = [panenTomat.tanggal_panen for panenTomat in allDataTomat]

    # Helper function to calculate percentage increase
    def calc_increase(data):
        if len(data) < 2 or 0 in data:
            return 0
        return round(((data[-1] - data[-2])/data[-2])*100)

    total_of_panen = sum(total_panen)
    totalPanenCabai = sum(stat_cabai)
    totalPanenTomat = sum(stat_tomat)

    if request.method == 'POST':
        kebun = request.form['kebun']
        komoditas = request.form['komoditas']
        jumlahBibit = request.form['jumlahBibit']
        tglBibit = request.form['tglBibit']

        add_data = DataPangan(kebun=kebun, komoditas=komoditas,
                              tanggal_bibit=tglBibit, jml_bibit=jumlahBibit,
                              status='Penanaman', jml_panen=0, tanggal_panen=0,
                              user_id=current_user.id)
        db.session.add(add_data)
        db.session.commit()
        flash('Berhasil menginput data!', 'success')
        return redirect(request.referrer)

    return render_template('dashboard/data-pangan.html', allDataCabai=allDataCabai, allDataTomat=allDataTomat, kelurahan=kel, user_data=user_data, kenaikan_cabai=calc_increase(stat_cabai), kenaikan_tomat=calc_increase(stat_tomat), stat_cabai=stat_cabai, stat_tomat=stat_tomat, cabai=cabai, tomat=tomat, pangan=pangan, total_panen=total_of_panen, totalPanenCabai=totalPanenCabai, totalPanenTomat=totalPanenTomat, tgl_panen_cabai=json.dumps(tgl_panen_cabai), tgl_panen_tomat=json.dumps(tgl_panen_tomat))

@views.route('/dashboard/data-pangan/import', methods=['GET', 'POST'])
@login_required
def import_data_pangan():
    if current_user.account_type == 'admin':
        return redirect(url_for('admin_page.index'))
    
    from openpyxl import load_workbook

    if request.method == 'POST':
        import_type = request.form['import_type']
        excel_file = request.files['excel_file'] 

        if 'excel_file' not in request.files:
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)

        if excel_file.filename == '':
            flash('Tidak ada file yang dipilih!', 'error')
            return redirect(request.url)

        # Validasi ekstensi dan nama file
        if excel_file and report_allowed_file(excel_file.filename):
            if not allowed_report_stat(excel_file.filename, import_type):
                flash('Nama file harus sesuai format ("panen" atau "penanaman") dan sesuai dengan pilihan status produksi!', 'warning')
                return redirect(request.url)

            filename = secure_filename(excel_file.filename)
            
            try: 
                wb = load_workbook(excel_file)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2):
                    kebun = row[0].value
                    komoditas = row[1].value
                    jml_bibit = row[2].value
                    tanggal_bibit = row[3].value.strftime('%Y-%m-%d') if isinstance(row[3].value, datetime) else row[3].value 

                    # Menyederhanakan pembuatan objek DataPangan berdasarkan tipe impor
                    if import_type == 'penanaman':
                        data_pangan = DataPangan(kebun=kebun, komoditas=komoditas, 
                                                jml_bibit=jml_bibit, tanggal_bibit=tanggal_bibit, 
                                                status='Penanaman', jml_panen=0, tanggal_panen=0, 
                                                user_id=current_user.id, kelurahan_id=current_user.kelurahan_id)
                    elif import_type == 'panen':
                        jml_panen = row[4].value 
                        tanggal_panen = row[5].value.strftime('%Y-%m-%d') if isinstance(row[5].value, datetime) else row[5].value

                        data_pangan = DataPangan(kebun=kebun, komoditas=komoditas, 
                                                jml_bibit=jml_bibit, tanggal_bibit=tanggal_bibit, 
                                                status='Panen', jml_panen=jml_panen, tanggal_panen=tanggal_panen, 
                                                user_id=current_user.id, kelurahan_id=current_user.kelurahan_id)
                    else:
                        flash('Tipe impor tidak valid!', 'error')
                        return redirect(url_for('views.import_data_pangan'))
                    db.session.add(data_pangan)
        
                db.session.commit()
                flash('Data berhasil diimpor!', 'success')
                return redirect(url_for('views.dataproduksi'))
            except Exception as e:
                flash(f'Terjadi kesalahan saat memproses file: {e}', 'error')
                return redirect(request.url)
        else:
            flash('Ekstensi file tidak diizinkan. Unggah file Excel (.xlsx)!', 'error')
            return redirect(request.url) 

    return render_template('dashboard/import_data.html')

@views.route('/dashboard/data-pangan/delete_selected', methods=['POST'])
@login_required
def delete_selected_data_pangan():
    delete_ids = request.form.getlist('delete_ids')  # Get list of selected IDs

    if delete_ids:
        DataPangan.query.filter(DataPangan.id.in_(delete_ids)).delete(synchronize_session=False)
        db.session.commit()
        flash('Data yang dipilih berhasil dihapus!', 'warning')
    else:
        flash('Tidak ada data yang dipilih!', 'warning')

    return redirect(url_for('views.dataproduksi'))

@views.route('/dashboard/data-pangan/update-data/<int:id>', methods=['POST', 'GET'])
@login_required
def updatepangan(id):
    pangan = DataPangan.query.get_or_404(id)
    kel = Kelurahan.query.filter_by(id=current_user.kelurahan_id).first()

    updateProd = request.form.get('updateProduksi')

    if updateProd == 'updateProduksi':
        if request.method == 'POST':
            # Memperbarui data pangan dengan data dari form
            pangan.kebun = request.form['updateKebun']
            pangan.komoditas = request.form['updateKomoditas']
            pangan.jml_bibit = request.form['updateJumlahBibit']
            pangan.tanggal_bibit = request.form['updateTglBibit']

            db.session.commit()
            return redirect(url_for('views.dataproduksi'))
    elif updateProd == 'dataPanen':
        if request.method == 'POST':
            # Memperbarui data pangan dan kelurahan dengan data dari form
            pangan.status = 'Panen'
            pangan.jml_panen = request.form['updateJumlahPanen']
            pangan.tanggal_panen = request.form['updateTglPanen']
            pangan.kelurahan_id = kel.id

            kel.jml_panen = request.form['updateJumlahPanen']

            db.session.commit()
            return redirect(request.referrer)

@views.route('/dashboard/data-pangan/delete-data/<int:id>', methods=['GET'])
@login_required
def delete_data_pangan(id):
    data = DataPangan.query.get_or_404(id)
    db.session.delete(data)
    db.session.commit()
    return redirect(url_for('views.dataproduksi'))

# todo ============== PROFILE PAGE ==============
@views.route('/dashboard/profil', methods=['GET', 'POST'])
@login_required
def profil():
    if current_user.account_type != 'user':
        return redirect(url_for('admin_page.index'))

    user = User.query.filter_by(id=current_user.id).first()
    kelurahan = Kelurahan.query.filter_by(id=user.kelurahan_id).first()

    return render_template('dashboard/profil.html', user=user, kelurahan=kelurahan)

@views.route('/dashboard/profil/<int:id>/update', methods=['GET', 'POST'])
# @login_required
def updateprofil(id):
    user = User.query.get_or_404(id)
    kelurahan = Kelurahan.query.filter_by(user_id=current_user.id).first()

    form_type = request.form.get('formType')

    if request.method == 'POST':
        if form_type == 'Data User':
            # Memperbarui data user dengan data dari form
            user.nama_lengkap = request.form['nama']
            user.username = request.form['username']
            user.pekerjaan = request.form['pekerjaan']
            user.kelamin = request.form['kelamin']
            user.bio = request.form['bio']

            if not kelurahan:
                add_kelurahan = Kelurahan(user_id=id)
                db.session.add(add_kelurahan)
            db.session.commit()
            flash('Profil Berhasil Diubah', 'success')
            return redirect(url_for('views.profil')) 

        elif form_type == 'Data Kelurahan':
            if kelurahan:  
                # Memperbarui data kelurahan dengan data dari form
                kelurahan.nama = request.form['kelurahan']
                kelurahan.kebun = request.form['kebun']
                kelurahan.luas_kebun = request.form['luaskebun']
                user.kelurahan_id = kelurahan.id
                db.session.commit()
                flash('Profil Berhasil diubah!', 'success')
            else:
                flash('Tidak ada data kelurahan untuk diperbaharui!', 'danger')
            return redirect(url_for('views.profil')) 

        else:
            flash('Tipe form tidak valid!', 'danger')
            return redirect(url_for('views.profil'))

    return render_template('dashboard/profil.html', user=user, kelurahan=kelurahan)

@views.route('/dashboard/pengaturan', methods=['GET', 'POST'])
@login_required
def settings():
    if current_user.account_type == 'admin':
        return redirect(url_for('admin_page.index'))

    user = User.query.filter_by(id=current_user.id).first()
    return render_template('dashboard/settings.html', user=user)

@views.route('/dashboard/pengaturan/<int:id>/update-email', methods=['GET', 'POST'])
@login_required
def updateemail(id):
    user = User.query.get_or_404(id)
    password = request.form['userPass']

    if request.method == 'POST':
        if check_password_hash(current_user.password, password):
            user.email = request.form['email']
            db.session.commit()
            flash('Email berhasil diubah!', category='success')
            return redirect((request.referrer))
        else:
            flash('Kata sandi salah, silakan coba lagi!', category='error')
            return redirect(url_for('views.settings')) 

@views.route('/dashboard/pengaturan/<int:id>/reset-profil', methods=['POST', 'GET'])
def resetprofil(id):
    kelurahan = Kelurahan.query.filter_by(user_id=id).first()

    if kelurahan:
        db.session.delete(kelurahan)
        db.session.add(Kelurahan(user_id=id))
        db.session.commit()
        flash('Profil anda berhasil direset!', 'warning')
        return redirect(url_for('views.profil'))
    else:
        flash('Tidak ada data kelurahan!', 'warning')
        return redirect(url_for('views.profil'))

@views.route('/dashboard/pengaturan/<int:id>/update-password', methods=['GET', 'POST'])
@login_required
def updatepassword(id):
    user = User.query.get_or_404(id)
    oldpass = request.form['old-pass']
    newpass = request.form['new-pass']
    confnewpass = request.form['new-pass-conf']

    if request.method == 'POST':
        if check_password_hash(current_user.password, oldpass):
            if confnewpass != newpass:
                flash('Konfirmasi kata sandi tidak cocok!', 'error')
            else:
                user.password = generate_password_hash(newpass, method='pbkdf2')
                db.session.commit()
                flash('Kata sandi berhasil diperbarui!', category='success')
                return redirect((request.referrer)) 
        else:
            flash('Kata sandi salah, silakan coba lagi!', category='error')
            return redirect(url_for('views.settings'))

@views.route('/prakiraan-cuaca', methods=['GET', 'POST'])
def weather():
    return render_template('weather.html')

# ========================= KELURAHAN SECTION =========================
@views.route('/peta-sebaran')
def mapbase():
    return render_template('kelurahan/map.html')

@views.route('/kelurahan-kulaba')
def kelkulaba():
    return render_template('kelurahan/kulaba.html')

@views.route('/kelurahan-sasa')
def kelsasa():
    return render_template('kelurahan/sasa.html')

@views.route('/kelurahan-kalumpang')
def kelkalumpang():
    return render_template('kelurahan/kalumpang.html')

@views.route('/kelurahan-santiong')
def kelsantiong():
    return render_template('kelurahan/santiong.html')

@views.route('/kelurahan-foramadiahi')
def kelforamadiahi():
    return render_template('kelurahan/foramadiahi.html')

@views.route('/kelurahan-tubo')
def keltubo():
    return render_template('kelurahan/tubo.html')

@views.route('/kelurahan-fitu')
def kelfitu():
    return render_template('kelurahan/fitu.html')

class MyHomeView(AdminIndexView):
    @expose('/')
    def index(self):
        user = User.query.all()
        return self.render('admin/index.html', user=user)

admin = Admin(index_view=MyHomeView())